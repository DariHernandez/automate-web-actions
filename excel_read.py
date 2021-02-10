#! python3

import openpyxl, os, random, pprint




class excel_file (): 
    """
    Read, write and process information from excel file
    """

    def __init__ (self, file_xlsx): 
        """
        Constructor of the class
        """

        current_path = os.path.dirname (__file__)
        self.__file_xlsx = os.path.join (current_path, "excel_files", file_xlsx)
        self.__wb = openpyxl.load_workbook (self.__file_xlsx)
    

    def get_data_sheet (self, sheet_name):
        """
        return all formated data from specific sheet of file
        """ 

        # Variable to return all data from sheet 
        data_return = []

        sheet = self.__wb[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column

        # Get headers of sheet
        column_names = []
        for column in range (1, max_column+1):
            column_name = sheet.cell (column = column, row = 1).value
            column_names.append (column_name)
        
        # Get data of cells
        for row in range (2, max_row+1): 
            data_row = {}
            for column in range (1, max_column+1): 
                cell_value = sheet.cell (column = column, row = row).value
                column_name = column_names[column-1]
                data_row[column_name] = cell_value
            data_return.append (data_row)

        return data_return

    def complite_random_sheet (self, sheet_name): 
        """
        complete spread sheet columns: email, phone, address, age, gender; 
        with first name, last name and random numbers

        Sorted columns of the sheet: first name, last name, email, phone, address, age, gender
        """

        wb = openpyxl.load_workbook (self.__file_xlsx)
        sheet = wb["sheet_name"]

        for row in range (1, 51): 

            # Get columns num
            first_name_column = 1
            last_name_column = 2
            email_column = 3
            phone_column = 4
            address_column = 5
            age_column = 6
            gender_column = 7
            
            # Get names
            first_name = sheet.cell (row=row, column=first_name_column).value
            last_name = sheet.cell (row=row, column=last_name_column).value
            
            # Create email
            email = first_name.lower() + last_name.lower() + "@gmail.com"

            # Create phone
            phone = "+41"
            for rand_num in range (0, 10): 
                number = random.randint(0,9)
                phone += str(number)
            
            # Create address
            street = random.randint (0,100)
            block = random.randint (0,100)
            address = "Street {}, block {}.".format (street, block)

            # Create age
            age = random.randint (18,90)

            # Crate gender
            gender = random.choices (["M", "F"])[0]
            
            # Write information
            sheet.cell (row=row, column=email_column).value = email
            sheet.cell (row=row, column=phone_column).value = phone
            sheet.cell (row=row, column=address_column).value = address
            sheet.cell (row=row, column=age_column).value = age
            sheet.cell (row=row, column=gender_column).value = gender



        wb.save(self.__file_xlsx)
        wb.close()